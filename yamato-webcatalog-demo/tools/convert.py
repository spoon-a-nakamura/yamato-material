# -*- coding: utf-8 -*-
"""仕様リスト_食品プラスチック(ボトル).xlsx → products.json 正規化スクリプト"""
import json, re, unicodedata, sys, os

import openpyxl

SRC = '/mnt/user-data/uploads/yamato_foods_catalog_pages/仕様リスト_食品プラスチック(ボトル).xlsx'
OUT_DIR = '/home/claude/work/data'
os.makedirs(OUT_DIR, exist_ok=True)

issues = []  # データ品質の記録（推測で埋めずに残す）

def z2h(s):
    if s is None: return None
    return unicodedata.normalize('NFKC', str(s)).replace('　', ' ').strip()

NUM = re.compile(r'-?\d+(?:\.\d+)?')

def num(v):
    """数値化。'1,000' / 'OF266' / '18・23' → 先頭の数値。取れなければ None"""
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = z2h(v).replace(',', '')
    m = NUM.search(s)
    return float(m.group()) if m else None

def clean_label(v):
    if v is None: return None
    s = z2h(v)
    s = re.sub(r'\s*\n\s*', ' ', s)
    return s or None

# ---------- 容量 ----------
def parse_capacity(raw, sku):
    lab = clean_label(raw)
    if lab is None:
        issues.append({'sku': sku, 'field': '容量', 'value': None, 'note': '空欄'})
        return None, 'unknown', None
    if lab.upper().startswith('=') or 'EXACT' in lab.upper():
        issues.append({'sku': sku, 'field': '容量', 'value': lab,
                       'note': 'Excel数式エラーが値として保存されている。原本の修正が必要'})
        return None, 'unknown', lab
    of = lab.upper().startswith('OF')
    n = num(lab)
    return n, ('overflow' if of else 'net'), lab

# ---------- 口部 ----------
MOUTH_TAGS = [
    ('F口(R)',  ['F口(R)', 'F口（R）']),
    ('F口',     ['F口']),
    ('26.3樹脂王冠', ['26.3樹脂王冠']),
    ('アルコア口', ['アルコア口']),
    ('専用キャップ', ['専用キャップ', '専用ネジキャップ']),
]
def parse_mouth(raw):
    lab = clean_label(raw)
    if not lab: return None, []
    tags = []
    rest = lab
    for tag, pats in MOUTH_TAGS:
        for p in pats:
            if p in rest:
                tags.append(tag)
                rest = rest.replace(p, '')
                break
    # F口(R) を拾ったら素の F口 は誤検出しないよう順序に依存
    return lab, tags

# ---------- 胴サイズ / 形状 ----------
SHAPE_WORDS = {'角': 'square', '楕円': 'oval', '六角': 'hex', '変形': 'freeform'}
def parse_body(raw, sku):
    lab = clean_label(raw)
    if not lab: return None, 'unknown', None, None
    shape = None
    for w, s in SHAPE_WORDS.items():
        if w in lab: shape = s; break
    nums = [float(x.replace(',', '')) for x in re.findall(r'\d+(?:\.\d+)?', lab.replace(',', ''))]
    w = d = None
    if lab.startswith('Φ') or lab.startswith('φ'):
        shape = shape or 'round'
        w = d = nums[0] if nums else None
    else:
        shape = shape or ('square' if len(nums) >= 2 else 'unknown')
        if len(nums) >= 2: w, d = nums[0], nums[1]
        elif len(nums) == 1: w = d = nums[0]
    if shape == 'hex' and len(nums) == 1:
        w = d = nums[0]
    return lab, shape, w, d

# ---------- アイコン ----------
ICON_MAP = [
    ('オリジナル', 'original'),
    ('バリア', 'barrier'),
    ('オーバーキャップ対応', 'overcap'),
    ('スクイズ', 'squeeze'),
    ('ラジェット', 'ratchet'),
]
def parse_icons(raw):
    lab = clean_label(raw)
    if not lab: return []
    out = []
    for word, key in ICON_MAP:
        if word in lab: out.append(key)
    return out

# ---------- リサイクルマーク（デモ用の派生項目） ----------
def parse_recycle(note):
    """マスターに独立列が無いため注釈から派生。提案書8-1では検索条件に必要な項目。"""
    if not note: return 'unknown'
    if 'PETマーク品あり' in note: return 'option'
    if 'PETマーク' in note: return 'yes'
    return 'unknown'

# ---------- 誌面ノンブル ----------
def page_for_sheet2(group, cap_ml, material, sku):
    g = group or ''
    if '調味料' in g:
        if cap_ml is None: return [], 'low'
        if cap_ml <= 280: return [29], 'high'
        if cap_ml <= 360: return [30], 'high'
        if cap_ml <= 870: return [31], 'high'
        return [32], 'high'
    if '広口' in g or '粉末' in g: return [33], 'high'
    if '単層' in g or 'チューブ' in g: return [34, 35], 'medium'
    if '密封' in g: return [36], 'high'
    if '飲料' in g:
        return ([37], 'high') if (material or '').startswith('PET') else ([38], 'high')
    return [], 'low'

SERIES_PAGE_RAW = {
    'Cシリーズ': ([24], 'high'),
    'FY＆AYシリーズ': ([25], 'high'),
    'ミニボトルシリーズ': ([26], 'medium'),
    'その他': ([27], 'medium'),
    '瓶口': ([27], 'low'),
}
SERIES_PAGE = {z2h(k): v for k, v in SERIES_PAGE_RAW.items()}

wb = openpyxl.load_workbook(SRC, data_only=True)
products = []

# ===== シート1: 調味料オリジナル =====
ws = wb['調味料オリジナル']
series = None
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    name = clean_label(row[2])
    if not name: continue
    if row[1]: series = clean_label(row[1])
    cap, cap_type, cap_label = parse_capacity(row[3], name)
    mouth_label, mouth_tags = parse_mouth(row[4])
    body_label, shape, bw, bd = parse_body(row[11], name)
    note = clean_label(row[10])
    pages, conf = SERIES_PAGE.get(z2h(series or ''), ([], 'unresolved'))
    if conf == 'unresolved':
        issues.append({'sku': name, 'field': '誌面ノンブル', 'value': series,
                       'note': 'シリーズと誌面ページの対応が支給資料から確定できない'})
    products.append(dict(
        sku=name, category='plastic-bottle', group='調味料（オリジナル）', series=series,
        variantGroup=None, isOriginal=True, isDemo=False,
        maker='ヤマトマテリアル（自社オリジナル）', makerUrl=None,
        capacityMl=cap, capacityType=cap_type, capacityLabel=cap_label,
        mouthLabel=mouth_label, mouthTypes=mouth_tags,
        weightG=num(row[5]), weightLabel=clean_label(row[5]),
        perCase=num(row[6]), perCaseLabel=clean_label(row[6]),
        material=clean_label(row[7]) or '—',
        heatResistC=num(row[8]), heatResistLabel=clean_label(row[8]) or '—',
        icons=parse_icons(row[9]), note=note,
        bodySizeLabel=body_label, shape=shape, bodyW=bw, bodyD=bd,
        heightMm=num(row[12]), labelHeightMm=num(row[13]),
        recycleMark=parse_recycle(note),
        catalogPages=pages, pageConfidence=conf,
        photoStatus=clean_label(row[14]),
    ))

# ===== シート2: 他（仕入れ品） =====
ws = wb['他']
group = None; variant = None
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    name = clean_label(row[3])
    if not name: continue
    if row[1]:
        group = clean_label(row[1]); variant = None
    if row[2]: variant = clean_label(row[2])
    if variant == 'キャップ':
        # キャップは別カテゴリー（誌面 65〜68）
        products.append(dict(
            sku=name, category='cap', group=group or 'キャップ', series=None,
            variantGroup='キャップ', isOriginal=False, isDemo=False,
            maker=clean_label(row[4]), makerUrl=None,
            capacityMl=None, capacityType='unknown', capacityLabel=None,
            mouthLabel=None, mouthTypes=[], weightG=None, weightLabel=None,
            perCase=None, perCaseLabel=None, material='—',
            heatResistC=None, heatResistLabel='—', icons=[], note=clean_label(row[12]),
            bodySizeLabel=None, shape='unknown', bodyW=None, bodyD=None,
            heightMm=None, labelHeightMm=None, recycleMark='unknown',
            catalogPages=[65], pageConfidence='low', photoStatus=clean_label(row[16]),
        ))
        continue
    cap, cap_type, cap_label = parse_capacity(row[5], name)
    mouth_label, mouth_tags = parse_mouth(row[6])
    material = clean_label(row[9]) or '—'
    body_label, shape, bw, bd = parse_body(row[13], name)
    note = clean_label(row[12])
    pages, conf = page_for_sheet2(group, cap, material, name)
    products.append(dict(
        sku=name, category='plastic-bottle', group=group or '—', series=None,
        variantGroup=variant if variant != 'キャップ' else None,
        isOriginal=False, isDemo=False,
        maker=clean_label(row[4]), makerUrl=None,
        capacityMl=cap, capacityType=cap_type, capacityLabel=cap_label,
        mouthLabel=mouth_label, mouthTypes=mouth_tags,
        weightG=num(row[7]), weightLabel=clean_label(row[7]),
        perCase=num(row[8]), perCaseLabel=clean_label(row[8]),
        material=material,
        heatResistC=num(row[10]), heatResistLabel=clean_label(row[10]) or '—',
        icons=parse_icons(row[11]), note=note,
        bodySizeLabel=body_label, shape=shape, bodyW=bw, bodyD=bd,
        heightMm=num(row[14]), labelHeightMm=num(row[15]),
        recycleMark=parse_recycle(note),
        catalogPages=pages, pageConfidence=conf,
        photoStatus=clean_label(row[16]),
    ))

# slug
seen = {}
for p in products:
    base = re.sub(r'[^0-9a-zA-Z]+', '-', z2h(p['sku'])).strip('-').lower() or 'sku'
    n = seen.get(base, 0); seen[base] = n + 1
    p['slug'] = base if n == 0 else f'{base}-{n+1}'

json.dump(products, open(f'{OUT_DIR}/products.real.json', 'w'), ensure_ascii=False, indent=1)
json.dump(issues, open(f'{OUT_DIR}/data-issues.json', 'w'), ensure_ascii=False, indent=1)

print('SKU数:', len(products))
from collections import Counter
print('カテゴリー:', Counter(p['category'] for p in products))
print('項目:', Counter(p['group'] for p in products))
print('形状:', Counter(p['shape'] for p in products))
print('口部:', Counter(t for p in products for t in p['mouthTypes']))
print('素材:', Counter(p['material'] for p in products))
print('リサイクル:', Counter(p['recycleMark'] for p in products))
print('ページ確度:', Counter(p['pageConfidence'] for p in products))
print('--- 検出した要確認事項', len(issues))
for i in issues[:20]: print('   ', i)
